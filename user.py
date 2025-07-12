import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, filedialog
import json
import os
import shutil
import pickle
from sklearn.preprocessing import LabelEncoder
from sklearn.svm import SVC
# from project.utils import Conf # Uncomment if you still use Conf class
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- Configuration des couleurs et polices (spécifiques à user.py) ---
COLOR_BG = "#F7FFF9"
COLOR_TEXT_DARK = "#325656"
COLOR_ACCENT = "#337D45" # New green for buttons and accents
COLOR_DANGER = "#DC3545"

FONT_TITLE = ("Segoe UI", 20, "bold")
FONT_HEADING = ("Segoe UI", 12, "bold")
FONT_NORMAL = ("Segoe UI", 10)
FONT_BUTTON = ("Segoe UI", 10, "bold")

# File paths - ensure these paths are correct
JSON_FILE_PATH_ENROLL = 'database/enroll.json'
JSON_FILE_PATH_ATTENDANCE = 'attendance.json'
DATASET_PATH = "dataset/PROJECT"

# Model paths (if Conf is not used)
ENCODINGS_PATH = "output/encodings.pickle"
RECOGNIZER_PATH = "output/recognizer.pickle"
LE_PATH = "output/le.pickle"

# If you use Conf, uncomment and adapt:
# try:
#     conf = Conf("config/config.json")
#     ENCODINGS_PATH = conf["encodings_path"]
#     RECOGNIZER_PATH = conf["recognizer_path"]
#     LE_PATH = conf["le_path"]
# except Exception as e:
#     # Fallback paths for development
#     ENCODINGS_PATH = "output/encodings.pickle"
#     RECOGNIZER_PATH = "output/recognizer.pickle"
#     LE_PATH = "output/le.pickle"
#     messagebox.showerror("Configuration Error", f"Failed to load configuration: {e}")


class UserManagementPage:
    def __init__(self, master_content_area, update_main_callback=None):
        """
        Initializes the UserManagementPage.

        Args:
            master_content_area (tk.Frame): The frame in main.py where this page will be displayed.
            update_main_callback (callable, optional): A callback function to notify main.py
                                                        of data changes (e.g., to refresh UI).
        """
        self.master_content_area = master_content_area
        self.update_main_callback = update_main_callback
        self.frame = tk.Frame(self.master_content_area, bg=COLOR_BG)
        self.frame.pack(fill="both", expand=True)

        self._configure_styles()
        self._create_widgets()
        self._load_and_display_enrollment_data()

    def _configure_styles(self):
        """Configures ttk widget styles for this page."""
        style = ttk.Style()
        
        def darken_color(hex_color, factor=0.8):
            hex_color = hex_color.lstrip('#')
            rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
            darker_rgb = tuple(int(c * factor) for c in rgb)
            return f'#{darker_rgb[0]:02x}{darker_rgb[1]:02x}{darker_rgb[2]:02x}'

        COLOR_ACCENT_DARK = darken_color(COLOR_ACCENT, 0.8)
        COLOR_ACCENT_DARKER = darken_color(COLOR_ACCENT, 0.6)
        COLOR_DANGER_DARK = darken_color(COLOR_DANGER, 0.8)
        COLOR_DANGER_DARKER = darken_color(COLOR_DANGER, 0.6)

        style.configure('TButton',
                        font=FONT_BUTTON,
                        foreground='white',
                        padding=10,
                        relief="flat")

        style.configure('Accent.TButton',
                        background=COLOR_ACCENT)
        style.map('Accent.TButton',
                  background=[('active', COLOR_ACCENT_DARK), ('pressed', COLOR_ACCENT_DARKER)],
                  foreground=[('active', 'white'), ('pressed', 'white')])

        style.configure('Danger.TButton',
                        background=COLOR_DANGER)
        style.map('Danger.TButton',
                  background=[('active', COLOR_DANGER_DARK), ('pressed', COLOR_DANGER_DARKER)],
                  foreground=[('active', 'white'), ('pressed', 'white')])
        
        style.configure("Treeview.Heading",
                        font=FONT_HEADING,
                        background=COLOR_ACCENT,
                        foreground="white",
                        relief="flat")
        style.map("Treeview.Heading",
                  background=[('active', COLOR_ACCENT_DARK)])
        
        style.configure("Treeview",
                        font=FONT_NORMAL,
                        rowheight=30,
                        background="white",
                        foreground=COLOR_TEXT_DARK,
                        fieldbackground="white",
                        bordercolor="#D3D3D3",
                        borderwidth=1)
        style.map("Treeview",
                  background=[('selected', '#C8E6C9')])

        style.configure("Vertical.TScrollbar",
                        background=COLOR_BG,
                        troughcolor=COLOR_BG,
                        gripcount=0,
                        gripcolor=COLOR_ACCENT,
                        bordercolor=COLOR_BG)
        style.map("Vertical.TScrollbar",
                  background=[('active', COLOR_TEXT_DARK)])
        
        style.configure('TMenubutton',
                        font=FONT_NORMAL,
                        background='white',
                        foreground=COLOR_TEXT_DARK,
                        relief="flat",
                        padding=(5, 5, 5, 5))
        style.map('TMenubutton',
                  background=[('active', '#E6E6E6')])

    def _create_widgets(self):
        """Creates the GUI widgets for the user management page using grid layout."""
        # Use grid for the main frame to allow precise placement and resizing
        # Configure rows and columns to expand as needed
        self.frame.grid_rowconfigure(0, weight=0) # Title row - fixed height
        self.frame.grid_rowconfigure(1, weight=1) # Treeview row - expands
        self.frame.grid_rowconfigure(2, weight=0) # Buttons row - fixed height
        self.frame.grid_columnconfigure(0, weight=1) # Single column - expands

        # Section Title
        section_title_frame = tk.Frame(self.frame, bg=COLOR_BG)
        section_title_frame.grid(row=0, column=0, sticky="ew", pady=(20, 15), padx=20)
        ttk.Label(section_title_frame, text="Enrolled Users Management", font=FONT_TITLE,
                  foreground=COLOR_TEXT_DARK, background=COLOR_BG).pack(anchor="w", padx=0)

        # Treeview for enrolled users
        tree_frame = tk.Frame(self.frame, bg=COLOR_BG)
        tree_frame.grid(row=1, column=0, sticky="nsew", pady=10, padx=20) # sticky "nsew" makes it fill the cell

        self.tree_enrollment = ttk.Treeview(tree_frame, columns=("ID", "Name", "Status"), show="headings")
        self.tree_enrollment.heading("ID", text="ID")
        self.tree_enrollment.heading("Name", text="Name")
        self.tree_enrollment.heading("Status", text="Status")
        self.tree_enrollment.column("ID", width=100, anchor="center")
        self.tree_enrollment.column("Name", width=250, anchor="center")
        self.tree_enrollment.column("Status", width=150, anchor="center")
        self.tree_enrollment.pack(side="left", fill="both", expand=True) # Treeview itself still uses pack within its frame

        # Scrollbar for Treeview
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_enrollment.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree_enrollment.config(yscrollcommand=scrollbar.set)

        # Buttons Frame - Now placed with grid
        buttons_frame = tk.Frame(self.frame, bg=COLOR_BG)
        buttons_frame.grid(row=2, column=0, sticky="ew", pady=(15, 20), padx=20) # sticky "ew" makes it fill horizontally

        # Apply buttons - Still use pack within the buttons_frame for horizontal alignment
        ttk.Button(buttons_frame, text="Delete User", command=self._delete_selected_user, style='Danger.TButton').pack(side="left", padx=5)
        ttk.Button(buttons_frame, text="Modify User", command=self._modify_selected_user, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(buttons_frame, text="Export to Excel", command=self._export_to_excel, style='Accent.TButton').pack(side="left", padx=5)
        ttk.Button(buttons_frame, text="Refresh List", command=self._load_and_display_enrollment_data, style='Accent.TButton').pack(side="right", padx=5)

    def _load_enrollment_data(self):
        try:
            with open(JSON_FILE_PATH_ENROLL, 'r') as file:
                enroll_data = json.load(file)
        except FileNotFoundError:
            enroll_data = {"student": {}}
        except json.JSONDecodeError:
            messagebox.showerror("File Error", f"Failed to decode JSON from {JSON_FILE_PATH_ENROLL}. File might be corrupted.")
            enroll_data = {"student": {}}
        return enroll_data

    def _load_attendance_data(self):
        try:
            with open(JSON_FILE_PATH_ATTENDANCE, 'r') as file:
                attendance_data = json.load(file)
        except FileNotFoundError:
            attendance_data = {"attendance": {}}
        except json.JSONDecodeError:
            messagebox.showerror("File Error", f"Failed to decode JSON from {JSON_FILE_PATH_ATTENDANCE}. File might be corrupted.")
            attendance_data = {"attendance": {}}
        return attendance_data

    def _load_and_display_enrollment_data(self):
        for item in self.tree_enrollment.get_children():
            self.tree_enrollment.delete(item)

        self.enroll_data = self._load_enrollment_data()
        self.enrollment_rows = []
        for primary_key, record_dict in self.enroll_data.get("student", {}).items():
            for person_id, details in record_dict.items():
                name = details[0] if len(details) > 0 else "unknown"
                status = details[1] if len(details) > 1 else "unknown"

                if (person_id and person_id.strip() != "" and person_id.lower() != "unknown" and
                    name and name.strip() != "" and name.lower() != "unknown" and
                    status and status.strip() != "" and status.lower() != "unknown"):
                    
                    self.enrollment_rows.append((person_id, name, status))
                    self.tree_enrollment.insert("", "end", values=(person_id, name, status))
                else:
                    print(f"Skipping incomplete or 'unknown' record: Primary Key '{primary_key}', Person ID '{person_id}', Name '{name}', Status '{status}'")

    def _delete_selected_user(self):
        selected_item = self.tree_enrollment.selection()
        if not selected_item:
            messagebox.showwarning("Selection Required", "Please select a user to delete.")
            return

        person_id = self.tree_enrollment.item(selected_item, 'values')[0]
        person_name = self.tree_enrollment.item(selected_item, 'values')[1]

        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Are you sure you want to delete {person_name} (ID: {person_id})?\n\nThis will permanently remove:\n- Enrollment record\n- Attendance records\n- Face images from dataset\n- Face encodings\n- And retrain the facial recognition model."
        )
        if not confirm:
            return

        try:
            enroll_data = self._load_enrollment_data()
            attendance_data = self._load_attendance_data()

            person_found_in_enroll = False
            primary_key_to_delete = None
            for pk, record_dict in enroll_data['student'].items():
                if person_id in record_dict:
                    primary_key_to_delete = pk
                    person_found_in_enroll = True
                    break

            if person_found_in_enroll:
                del enroll_data['student'][primary_key_to_delete]
            else:
                messagebox.showerror("Deletion Error", f"No enrollment record found for ID {person_id}.")
                return

            user_dataset_path = os.path.join(DATASET_PATH, person_id)
            if os.path.exists(user_dataset_path):
                shutil.rmtree(user_dataset_path)
            
            if person_id in attendance_data["attendance"]:
                del attendance_data["attendance"][person_id]

            try:
                with open(ENCODINGS_PATH, "rb") as f:
                    data = pickle.load(f)
            except FileNotFoundError:
                messagebox.showwarning("Encodings Missing", "Encodings file not found. Skipping model retraining.")
                data = {"names": [], "encodings": []}

            if person_id in data['names']:
                indices_to_delete = [i for i, name in enumerate(data['names']) if name == person_id]
                for index in reversed(indices_to_delete):
                    del data['names'][index]
                    del data['encodings'][index]

                with open(ENCODINGS_PATH, "wb") as f:
                    pickle.dump(data, f)
                
                unique_names = list(set(data['names']))
                if len(unique_names) >= 2:
                    le = LabelEncoder()
                    labels = le.fit_transform(data["names"])
                    recognizer = SVC(C=1.0, kernel="linear", probability=True)
                    recognizer.fit(data["encodings"], labels)

                    with open(RECOGNIZER_PATH, "wb") as f:
                        pickle.dump(recognizer, f)
                    with open(LE_PATH, "wb") as f:
                        pickle.dump(le, f)
                else:
                    if os.path.exists(RECOGNIZER_PATH): os.remove(RECOGNIZER_PATH)
                    if os.path.exists(LE_PATH): os.remove(LE_PATH)
                    messagebox.showinfo("Model Update", "Not enough persons to retrain model, or model/LabelEncoder removed.")

            with open(JSON_FILE_PATH_ENROLL, 'w') as file:
                json.dump(enroll_data, file, indent=4)
            with open(JSON_FILE_PATH_ATTENDANCE, 'w') as file:
                json.dump(attendance_data, file, indent=4)

            messagebox.showinfo("Success", f"User {person_name} (ID: {person_id}) deleted successfully.")
            self._load_and_display_enrollment_data()
            if self.update_main_callback:
                self.update_main_callback()

        except Exception as e:
            messagebox.showerror("Operation Failed", f"An error occurred during deletion: {e}")

    def _modify_selected_user(self):
        selected_item = self.tree_enrollment.selection()
        if not selected_item:
            messagebox.showwarning("Selection Required", "Please select a user to modify.")
            return

        current_id, current_name, current_status = self.tree_enrollment.item(selected_item, 'values')

        modify_window = tk.Toplevel(self.master_content_area.winfo_toplevel())
        modify_window.title(f"Modify User: {current_name}")
        modify_window.transient(self.master_content_area.winfo_toplevel())
        modify_window.grab_set()
        
        main_window = self.master_content_area.winfo_toplevel()
        main_window.update_idletasks()
        main_x = main_window.winfo_x()
        main_y = main_window.winfo_y()
        main_width = main_window.winfo_width()
        main_height = main_window.winfo_height()

        window_width = 450
        window_height = 220
        x_pos = main_x + (main_width // 2) - (window_width // 2)
        y_pos = main_y + (main_height // 2) - (window_height // 2)
        
        modify_window.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")
        modify_window.resizable(False, False)
        modify_window.config(bg=COLOR_BG)

        content_frame = tk.Frame(modify_window, bg=COLOR_BG, padx=25, pady=25)
        content_frame.pack(expand=True, fill="both")

        tk.Label(content_frame, text="ID:", font=FONT_NORMAL, bg=COLOR_BG, fg=COLOR_TEXT_DARK).grid(row=0, column=0, padx=10, pady=7, sticky="w")
        id_entry = tk.Entry(content_frame, font=FONT_NORMAL, bg="white", fg=COLOR_TEXT_DARK, relief="flat", bd=1)
        id_entry.insert(0, current_id)
        id_entry.config(state="readonly")
        id_entry.grid(row=0, column=1, padx=10, pady=7, sticky="ew")

        tk.Label(content_frame, text="Name:", font=FONT_NORMAL, bg=COLOR_BG, fg=COLOR_TEXT_DARK).grid(row=1, column=0, padx=10, pady=7, sticky="w")
        name_entry = tk.Entry(content_frame, font=FONT_NORMAL, bg="white", fg=COLOR_TEXT_DARK, relief="flat", bd=1)
        name_entry.insert(0, current_name)
        name_entry.grid(row=1, column=1, padx=10, pady=7, sticky="ew")

        tk.Label(content_frame, text="Status:", font=FONT_NORMAL, bg=COLOR_BG, fg=COLOR_TEXT_DARK).grid(row=2, column=0, padx=10, pady=7, sticky="w")
        status_var = tk.StringVar(content_frame)
        status_var.set(current_status)
        status_options = ["active", "inactive"]
        
        status_menu = ttk.OptionMenu(content_frame, status_var, current_status, *status_options)
        status_menu.grid(row=2, column=1, padx=10, pady=7, sticky="ew")
        
        content_frame.grid_columnconfigure(1, weight=1)

        def save_modifications():
            new_name = name_entry.get().strip()
            new_status = status_var.get()

            if not new_name:
                messagebox.showwarning("Input Error", "Name cannot be empty. Please enter a valid name.")
                return

            enroll_data = self._load_enrollment_data()
            person_modified = False

            for primary_id, record_dict in enroll_data['student'].items():
                if current_id in record_dict:
                    record_dict[current_id][0] = new_name
                    record_dict[current_id][1] = new_status
                    person_modified = True
                    break

            if person_modified:
                try:
                    with open(JSON_FILE_PATH_ENROLL, 'w') as file:
                        json.dump(enroll_data, file, indent=4)
                    messagebox.showinfo("Success", f"User '{current_name}' (ID: {current_id}) modified successfully to '{new_name}'.")
                    self._load_and_display_enrollment_data()
                    if self.update_main_callback:
                        self.update_main_callback()
                    modify_window.destroy()
                except Exception as e:
                    messagebox.showerror("Save Error", f"Failed to save modifications: {e}")
            else:
                messagebox.showerror("Error", "Could not find user record to modify.")

        save_button = ttk.Button(content_frame, text="Save Changes", command=save_modifications, style='Accent.TButton')
        save_button.grid(row=3, column=0, columnspan=2, pady=15)

        content_frame.grid_rowconfigure(len(content_frame.grid_slaves(column=0)) + 1, weight=1)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_columnconfigure(1, weight=1)

    def _export_to_excel(self):
        if not hasattr(self, 'enrollment_rows') or not self.enrollment_rows:
            messagebox.showinfo("No Data", "No enrolled users to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="enrolled_users.xlsx"
        )
        if not file_path:
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Enrolled Users"

            headers = ["ID", "Name", "Status"]
            sheet.append(headers)

            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color=COLOR_ACCENT.lstrip('#'), end_color=COLOR_ACCENT.lstrip('#'), fill_type="solid")
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = header_fill
                thin_border = Border(left=Side(style='thin', color="A0A0A0"),
                                     right=Side(style='thin', color="A0A0A0"),
                                     top=Side(style='thin', color="A0A0A0"),
                                     bottom=Side(style='thin', color="A0A0A0"))
                cell.border = thin_border

            for row_data in self.enrollment_rows:
                sheet.append(row_data)
                for cell in sheet[sheet.max_row]:
                    thin_border = Border(left=Side(style='thin', color="D3D3D3"),
                                         right=Side(style='thin', color="D3D3D3"),
                                         top=Side(style='thin', color="D3D3D3"),
                                         bottom=Side(style='thin', color="D3D3D3"))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Export Successful", f"Enrolled users exported to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export data to Excel: {e}")


# Example of how to use it if running user.py directly for testing:
if __name__ == "__main__":
    root = tk.Tk()
    root.title("User Management Test (user.py standalone)")
    root.geometry("900x600") # You can adjust this to see how it affects sizing
    root.config(bg=COLOR_BG)

    test_content_area = tk.Frame(root, bg=COLOR_BG)
    test_content_area.pack(fill="both", expand=True, padx=20, pady=20)

    user_manager = UserManagementPage(test_content_area)

    root.mainloop()